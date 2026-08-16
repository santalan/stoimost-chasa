from __future__ import annotations
import re, hashlib
from pathlib import Path
from collections import defaultdict
from datetime import date, datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook, Workbook
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib.dates as mdates

APP='Стоимость часа'; VERSION='0.2.0'
NORM={
2021:[90,113,132,131,114,125,132,132,132,126,119,132],2022:[96,113,131,126,108,126,126,138,132,126,125,132],
2023:[102,107,131,120,120,126,126,138,126,132,125,126],2024:[102,119,119,126,119,113,138,132,126,138,125,126],
2025:[102,120,125,131,108,113,138,126,132,138,113,132],2026:[90,114,126,131,113,125,138,126,132,132,119,132]}
STD=sum(map(sum,NORM.values()))/len(NORM)/12
CATS={'stim':'Стимулирующие (ПДД, ВМП)','bonus':'Премии','vacation':'Отпускные','expansion':'Расширение','night':'Ночные','holiday':'Праздничные / выходные','sick':'Больничные','comp':'Компенсации','other':'Прочие начисления'}
MODES={'as_is':'Как есть','spread':'Распределить','exclude':'Не учитывать'}
DEFAULT={'stim':'spread','bonus':'exclude','vacation':'spread','expansion':'as_is','night':'as_is','holiday':'as_is','sick':'exclude','comp':'exclude','other':'as_is'}
CPI21={1:100.67,2:100.78,3:100.66,4:100.58,5:100.74,6:100.69,7:100.31,8:100.17,9:100.60,10:101.11,11:100.96,12:100.82}
YOY={2022:[8.73,9.15,16.69,17.83,17.10,15.90,15.10,14.30,13.68,12.63,11.98,11.94],2023:[11.77,10.99,3.51,2.31,2.51,3.25,4.30,5.15,6.00,6.69,7.48,7.42],2024:[7.44,7.69,7.72,7.84,8.30,8.59,9.13,9.05,8.63,8.54,8.88,9.52],2025:[9.92,10.06,10.34,10.23,9.88,9.40,8.79,8.14,7.98,7.71,6.64,5.59],2026:[6.00,5.91,5.86,5.58,5.31,6.02,5.98]}
RATE={2021:[4.25,4.25,4.50,5,5,5.5,6.5,6.5,6.75,7.5,7.5,8.5],2022:[8.5,20,20,17,11,9.5,8,8,7.5,7.5,7.5,7.5],2023:[7.5,7.5,7.5,7.5,7.5,7.5,8.5,12,13,15,15,16],2024:[16,16,16,16,16,16,18,18,19,21,21,21],2025:[21,21,21,21,21,20,18,18,17,16.5,16.5,16],2026:[16,15.5,15,14.5,14.5,14.25,14]}

def fnum(v):
    if isinstance(v,(int,float)): return float(v)
    try:return float(str(v).replace(' ','').replace(',','.'))
    except:return None

def period(v):
    m=re.fullmatch(r'(\d{1,2})-(\d{4})',str(v or '').strip()); return (int(m.group(2)),int(m.group(1))) if m else None

def classify(s):
    x=s.lower().strip()
    if 'отпуск' in x and 'комп' not in x:return 'vacation'
    if 'больнич' in x or x.startswith('бл '):return 'sick'
    if 'прем' in x:return 'bonus'
    if 'расшир' in x:return 'expansion'
    if 'ночн' in x:return 'night'
    if 'празд' in x or 'выходн' in x:return 'holiday'
    if 'стим' in x or 'пдд' in x or 'вмп' in x:return 'stim'
    if 'комп' in x or 'задерж' in x:return 'comp'
    if x.startswith('оклад') or 'вредност' in x or 'выслуг' in x:return 'base'
    return 'other'

def parse_files(paths):
    out=[]; seen=set(); warnings=[]
    for p in paths:
        h=hashlib.sha256(Path(p).read_bytes()).hexdigest()
        if h in seen:continue
        seen.add(h)
        try: ws=load_workbook(p,data_only=True,read_only=True).active
        except Exception as e: warnings.append(f'{Path(p).name}: {e}'); continue
        role=False
        for row in ws.iter_rows(values_only=True):
            b=str(row[1] or '').strip() if len(row)>1 else ''
            if b=='Должность:':
                txt=' '.join(str(x or '') for x in row[4:7]); role='Врач-рентгенолог' in txt; continue
            if role and b=='Всего начислено': role=False; continue
            if not role or not b: continue
            per=period(row[5] if len(row)>5 else None); amt=fnum(row[14] if len(row)>14 else None)
            if not per or amt is None:continue
            hrs=0.0
            if b.lower().startswith('оклад'):
                for idx in (10,9):
                    if len(row)>idx:
                        nums=re.findall(r'-?\d+(?:[.,]\d+)?',str(row[idx] or ''))
                        if nums: hrs=float(nums[-1].replace(',','.')); break
            salary=None
            if b.lower()=='оклад' and hrs>0:
                norm=NORM.get(per[0],[None]*12)[per[1]-1]
                if norm and hrs>0: salary=amt*norm/hrs
            out.append({'y':per[0],'m':per[1],'label':b,'amt':amt,'hours':hrs,'cat':classify(b),'file':Path(p).name,'salary':salary})
    return out,warnings

def detect_salary(accr, months):
    vals=defaultdict(list)
    for a in accr:
        if a['salary'] and 10000<a['salary']<300000: vals[(a['y'],a['m'])].append(a['salary'])
    raw={k:(sorted(v)[len(v)//2] if v else None) for k,v in vals.items()}
    levels=[]; last=None
    for k in months:
        v=raw.get(k)
        if v:
            candidates=[40000,48000,90000,115000]
            near=min(candidates,key=lambda x:abs(x-v))
            v=near if abs(near-v)/near<0.08 else round(v/1000)*1000
            last=v
        levels.append(last)
    first=next((x for x in levels if x),40000)
    levels=[x or first for x in levels]
    pid=[]; n=0; prev=None
    for v in levels:
        if prev is None or abs(v-prev)>1: n+=1
        pid.append(n); prev=v
    return levels,pid

def analyze(accr,modes):
    keys=sorted({(a['y'],a['m']) for a in accr})
    if not keys:return [],[]
    y,m=min(keys); end=max(keys); months=[]
    while (y,m)<=end:
        months.append((y,m)); m+=1
        if m==13:y+=1;m=1
    levels,pids=detect_salary(accr,months)
    by=defaultdict(lambda:defaultdict(float)); hrs=defaultdict(float)
    for a in accr:
        by[(a['y'],a['m'])][a['cat']]+=a['amt']; hrs[(a['y'],a['m'])]+=a['hours']
    rows=[]
    for i,(y,m) in enumerate(months):
        norm=NORM.get(y,[0]*12)[m-1]; h=hrs[(y,m)]; std=(h/norm*STD) if norm and h>0 else 0
        rows.append({'date':datetime(y,m,1),'y':y,'m':m,'norm':norm,'hours':h,'std':std,'salary':levels[i],'pid':pids[i],'cats':dict(by[(y,m)])})
    allocated=[defaultdict(float) for _ in rows]
    for cat in ['base']+list(CATS):
        mode='as_is' if cat=='base' else modes.get(cat,DEFAULT.get(cat,'as_is'))
        if mode=='exclude':continue
        if mode=='as_is':
            for i,r in enumerate(rows): allocated[i][cat]=r['cats'].get(cat,0)
        else:
            groups=defaultdict(list)
            for i,r in enumerate(rows): groups[r['y'] if cat=='vacation' else r['pid']].append(i)
            for _,idxs in groups.items():
                pool=sum(rows[i]['cats'].get(cat,0) for i in idxs); den=sum(rows[i]['std'] for i in idxs)
                if den:
                    for i in idxs: allocated[i][cat]=pool*rows[i]['std']/den
    for i,r in enumerate(rows):
        r['alloc']=dict(allocated[i]); r['total']=sum(allocated[i].values()); r['rate']=r['total']/r['std'] if r['std'] else None
    full=[r for r in rows if r['rate'] is not None]
    if full and full[0]['y']==2021:
        firsty=full[0]['y']; first=[r for r in full if r['y']==firsty]; base=sum(r['total'] for r in first)/sum(r['std'] for r in first)
        pl={}; lv=1
        for mo in range(1,13): lv*=CPI21[mo]/100; pl[(2021,mo)]=lv
        for yr in range(2022,2027):
            for mo in range(1,len(YOY.get(yr,[]))+1): pl[(yr,mo)]=pl[(yr-1,mo)]*(1+YOY[yr][mo-1]/100)
        sp=pl[(2021,full[0]['m'])]; inv=base
        for r in rows:
            k=(r['y'],r['m']); r['infl']=base*pl[k]/sp if k in pl else None
            if k==(2021,full[0]['m']): inv=base
            elif r['y'] in RATE and r['m']<=len(RATE[r['y']]): inv*=1+RATE[r['y']][r['m']-1]/1200
            r['cbr']=inv
    return rows,[(p,levels[i]) for i,p in enumerate(pids)]

class App(tk.Tk):
    def __init__(self):
        super().__init__(); self.title(f'{APP} {VERSION}'); self.geometry('1280x800'); self.minsize(1000,680)
        self.paths=[]; self.accr=[]; self.rows=[]; self.mode={k:tk.StringVar(value=DEFAULT[k]) for k in CATS}; self.stacked=tk.BooleanVar(False); self.infl=tk.BooleanVar(True); self.cbr=tk.BooleanVar(True)
        self.build()
    def build(self):
        root=ttk.Frame(self,padding=10); root.pack(fill='both',expand=True)
        ttk.Label(root,text='Стоимость часа',font=('Segoe UI',18,'bold')).pack(anchor='w')
        ttk.Label(root,text='Расчётные листки врача-рентгенолога · фиксированная 30-часовая неделя').pack(anchor='w',pady=(0,8))
        top=ttk.Frame(root); top.pack(fill='x')
        left=ttk.LabelFrame(top,text='1. Расчётные листки',padding=8); left.pack(side='left',fill='both',expand=True,padx=(0,5))
        b=ttk.Frame(left); b.pack(fill='x'); ttk.Button(b,text='Добавить файлы…',command=self.add_files).pack(side='left'); ttk.Button(b,text='Добавить папку…',command=self.add_folder).pack(side='left',padx=5); ttk.Button(b,text='Очистить',command=self.clear).pack(side='left')
        self.list=tk.Listbox(left,height=8); self.list.pack(fill='both',expand=True,pady=(5,0))
        right=ttk.LabelFrame(top,text='2. Как учитывать выплаты',padding=8); right.pack(side='left',fill='both',expand=True,padx=(5,0))
        reverse={v:k for k,v in MODES.items()}
        for i,(k,label) in enumerate(CATS.items()):
            ttk.Label(right,text=label).grid(row=i,column=0,sticky='w',padx=(0,8),pady=2)
            cb=ttk.Combobox(right,state='readonly',values=list(MODES.values()),width=18); cb.set(MODES[self.mode[k].get()]); cb.grid(row=i,column=1,sticky='w')
            cb.bind('<<ComboboxSelected>>',lambda e,key=k,w=cb:self.mode[key].set(reverse[w.get()]))
        act=ttk.Frame(root); act.pack(fill='x',pady=8)
        ttk.Button(act,text='Построить график',command=self.calc).pack(side='left'); ttk.Button(act,text='Экспорт PNG…',command=self.png).pack(side='left',padx=5); ttk.Button(act,text='Экспорт Excel…',command=self.xlsx).pack(side='left')
        ttk.Checkbutton(act,text='Stacked',variable=self.stacked,command=self.draw).pack(side='right'); ttk.Checkbutton(act,text='Ставка ЦБ',variable=self.cbr,command=self.draw).pack(side='right',padx=5); ttk.Checkbutton(act,text='Инфляция',variable=self.infl,command=self.draw).pack(side='right')
        self.fig=Figure(figsize=(11,5.5),dpi=100); self.ax=self.fig.add_subplot(111); self.canvas=FigureCanvasTkAgg(self.fig,master=root); self.canvas.get_tk_widget().pack(fill='both',expand=True); NavigationToolbar2Tk(self.canvas,root).update()
        self.status=ttk.Label(root,text='Добавьте расчётные листки'); self.status.pack(anchor='w',pady=(5,0))
    def add_files(self):
        f=filedialog.askopenfilenames(filetypes=[('Excel','*.xlsx')]); self.add(f)
    def add_folder(self):
        d=filedialog.askdirectory(); self.add(sorted(str(x) for x in Path(d).glob('*.xlsx')) if d else [])
    def add(self,paths):
        for p in paths:
            if p not in self.paths:self.paths.append(p)
        self.list.delete(0,'end'); [self.list.insert('end',Path(p).name) for p in self.paths]; self.status.config(text=f'Выбрано файлов: {len(self.paths)}')
    def clear(self):self.paths=[];self.accr=[];self.rows=[];self.list.delete(0,'end');self.draw()
    def calc(self):
        if not self.paths:return messagebox.showinfo(APP,'Сначала добавьте расчётные листки.')
        self.accr,w=parse_files(self.paths); self.rows,_=analyze(self.accr,{k:v.get() for k,v in self.mode.items()}); self.status.config(text=f'Начислений: {len(self.accr)}. '+('Предупреждения: '+str(len(w)) if w else 'Ошибок чтения нет.')); self.draw()
    def draw(self):
        self.ax.clear()
        if not self.rows:self.canvas.draw();return
        d=[r['date'] for r in self.rows]; valid=[r for r in self.rows if r['rate'] is not None]
        if self.stacked.get() and valid:
            cats=['base']+[k for k in CATS if self.mode[k].get()!='exclude']; labels=['Базовая часть']+[CATS[k] for k in cats[1:]]
            self.ax.stackplot([r['date'] for r in valid],*[[r['alloc'].get(c,0)/(r['std'] or 1) for r in valid] for c in cats],labels=labels,alpha=.8)
            self.ax.plot([r['date'] for r in valid],[r['rate'] for r in valid],label='Итого',linewidth=1.5)
        else:self.ax.plot(d,[r['rate'] for r in self.rows],marker='o',markersize=3,label='Стоимость часа')
        if self.infl.get() and any(r.get('infl') for r in self.rows):self.ax.plot(d,[r.get('infl') for r in self.rows],label='Инфляционный ориентир')
        if self.cbr.get() and any(r.get('cbr') for r in self.rows):self.ax.plot(d,[r.get('cbr') for r in self.rows],label='Ориентир по ставке ЦБ')
        self.ax.set_ylim(bottom=0); self.ax.set_ylabel('₽/ч'); self.ax.grid(True,alpha=.25); self.ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3)); self.ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m')); self.fig.autofmt_xdate(); self.ax.legend(loc='upper left'); self.fig.tight_layout(); self.canvas.draw()
    def png(self):
        if not self.rows:return
        p=filedialog.asksaveasfilename(defaultextension='.png',filetypes=[('PNG','*.png')]);
        if p:self.fig.savefig(p,dpi=200,bbox_inches='tight')
    def xlsx(self):
        if not self.rows:return
        p=filedialog.asksaveasfilename(defaultextension='.xlsx',filetypes=[('Excel','*.xlsx')]);
        if not p:return
        wb=Workbook(); ws=wb.active; ws.title='Помесячно'; cats=['base']+list(CATS)
        ws.append(['Месяц','Норма 30ч','Факт. часы','Нормализованные часы','Оклад']+[('Базовая часть' if c=='base' else CATS[c])+', ₽' for c in cats]+['Итого, ₽','Стоимость часа, ₽/ч'])
        for r in self.rows:ws.append([date(r['y'],r['m'],1),r['norm'],r['hours'],r['std'],r['salary']]+[r['alloc'].get(c,0) for c in cats]+[r['total'],r['rate']])
        ac=wb.create_sheet('Начисления'); ac.append(['Файл','Период','Начисление','Сумма','Часы','Категория'])
        for a in self.accr:ac.append([a['file'],f"{a['m']:02d}.{a['y']}",a['label'],a['amt'],a['hours'],a['cat']])
        st=wb.create_sheet('Настройки'); st.append(['Категория','Режим']); [st.append([CATS[k],MODES[v.get()]]) for k,v in self.mode.items()]; wb.save(p)

def main():App().mainloop()
if __name__=='__main__':main()
